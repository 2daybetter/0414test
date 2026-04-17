"""
gen_requirements.py — 요구사항정의서 Excel 생성기 (AY-01)
Usage: python gen_requirements.py <data.json> <output.xlsx>
"""
import json, sys, os
from openpyxl import Workbook
sys.path.insert(0, os.path.dirname(__file__))
from adeo_style import *

PRIORITY_COLORS = {
    "Must":   ("FDECEA", "C62828"),
    "Should": ("FFF8E1", "F57F17"),
    "Nice":   ("E8F5E9", "2E7D32"),
}
TYPE_COLORS = {
    "FO": ("D6E4F0", C_NAVY),
    "BO": ("E8F3E8", "1B5E20"),
}


def build_cover(wb, d):
    ws = wb.active
    ws.title = "문서 정보"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [3, 22, 32, 18, 3]):
        set_col_width(ws, col, w)

    fill_range(ws, 1, 1, 3, 5, FILL_NAVY)
    merge_write(ws, 2, 2, 2, 3, "ADEO GROUP", font=fnt(18, True, C_WHITE), fill=FILL_NAVY, alignment=ALIGN_LEFT)
    merge_write(ws, 2, 4, 2, 5, "요구사항정의서 AY-01", font=fnt(11, False, "A0C4E8"), fill=FILL_NAVY,
                alignment=Alignment(horizontal="right", vertical="center"))
    for r in range(1, 4):
        set_row_height(ws, r, 22)

    merge_write(ws, 5, 2, 6, 4, d["project_name"],
                font=fnt(18, True, C_NAVY), fill=FILL_WHITE,
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
    set_row_height(ws, 5, 28); set_row_height(ws, 6, 28)
    merge_write(ws, 7, 2, 7, 4, "", fill=fill(C_BLUE))
    set_row_height(ws, 7, 4)

    info = [
        ("문서 코드", "AY-01"), ("버전", d.get("version", "v1.0")),
        ("작성일", d.get("doc_date", "")), ("고객사", d.get("client_name", "")),
        ("작성자", d.get("author", "웹기획팀")), ("검토자", d.get("reviewer", "이사님")),
    ]
    for i, (k, v) in enumerate(info, 9):
        write_cell(ws, i, 2, k, font=FONT_BOLD, fill=FILL_BLUE_LT, alignment=ALIGN_CENTER, border=BORDER_ALL)
        merge_write(ws, i, 3, i, 4, v, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_LEFT, border=BORDER_ALL)
        set_row_height(ws, i, 20)


def build_analysis(wb, d):
    ws = wb.create_sheet("현황 분석")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [3, 22, 50, 30, 3]):
        set_col_width(ws, col, w)

    section_title(ws, 1, 4, "1. 현황 분석 (As-Is / To-Be)", height=30)

    sub_section(ws, 3, 4, "  1-1. 고객사 현황 (As-Is)", height=20)
    header_row(ws, 4, ["", "분석 항목", "현황 내용", "개선 필요 사항", ""], height=22)
    for i, item in enumerate(d.get("as_is", []), 1):
        row = 4 + i
        alt = i % 2 == 0
        bg = FILL_ACCENT if alt else FILL_GRAY
        write_cell(ws, row, 2, item["category"], font=FONT_BOLD, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 3, item["current"],  font=FONT_BODY, fill=bg, alignment=ALIGN_LEFT_TOP, border=BORDER_ALL)
        write_cell(ws, row, 4, item["issue"],    font=FONT_BODY, fill=bg, alignment=ALIGN_LEFT_TOP, border=BORDER_ALL)
        set_row_height(ws, row, 40)

    hr = 5 + len(d.get("as_is", []))
    sub_section(ws, hr, 4, "  1-2. 목표 방향 (To-Be)", height=20)
    header_row(ws, hr + 1, ["", "목표 항목", "목표 내용", "기대 효과", ""], height=22)
    for i, item in enumerate(d.get("to_be", []), 1):
        row = hr + 1 + i
        alt = i % 2 == 0
        bg = FILL_ACCENT if alt else FILL_GRAY
        write_cell(ws, row, 2, item["goal"],   font=FONT_BOLD, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 3, item["detail"], font=FONT_BODY, fill=bg, alignment=ALIGN_LEFT_TOP, border=BORDER_ALL)
        write_cell(ws, row, 4, item["effect"], font=FONT_BODY, fill=bg, alignment=ALIGN_LEFT_TOP, border=BORDER_ALL)
        set_row_height(ws, row, 40)

    # 벤치마크
    br = hr + 3 + len(d.get("to_be", []))
    sub_section(ws, br, 4, "  1-3. 벤치마크 사이트", height=20)
    header_row(ws, br + 1, ["", "사이트명", "URL", "참고 포인트", ""], height=22)
    for i, b in enumerate(d.get("benchmarks", []), 1):
        data_row(ws, br + 1 + i, ["", b["name"], b["url"], b["point"], ""], i % 2 == 0)


def build_fo_req(wb, d):
    ws = wb.create_sheet("FO 요구사항")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [3, 12, 18, 50, 14, 20, 18, 3]):
        set_col_width(ws, col, w)

    section_title(ws, 1, 7, "2. FO (Front Office) 기능 요구사항", height=30)
    header_row(ws, 3,
               ["", "요구사항 ID", "구분", "요구사항 내용", "우선순위", "관련 화면", "비고", ""],
               height=22)

    for i, req in enumerate(d.get("fo_requirements", []), 1):
        row = 3 + i
        pri = req.get("priority", "Must")
        pc, fc = PRIORITY_COLORS.get(pri, (C_GRAY, "000000"))
        bg = fill(pc)
        write_cell(ws, row, 2, req["id"],       font=FONT_SMALL, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 3, req["category"], font=FONT_BODY,  fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 4, req["content"],  font=FONT_BODY,  fill=bg, alignment=ALIGN_LEFT_TOP, border=BORDER_ALL)
        write_cell(ws, row, 5, pri, font=fnt(10, True, fc), fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 6, req.get("screen", ""), font=FONT_SMALL, fill=bg, alignment=ALIGN_LEFT, border=BORDER_ALL)
        write_cell(ws, row, 7, req.get("note", ""),   font=FONT_SMALL, fill=bg, alignment=ALIGN_LEFT, border=BORDER_ALL)
        set_row_height(ws, row, 36)

    # 범례
    lr = 4 + len(d.get("fo_requirements", []))
    sub_section(ws, lr, 7, "  우선순위 범례", height=18)
    for j, (p, (pc, fc)) in enumerate(PRIORITY_COLORS.items(), 1):
        write_cell(ws, lr + 1, j + 1, p, font=fnt(10, True, fc), fill=fill(pc), alignment=ALIGN_CENTER, border=BORDER_ALL)
    set_row_height(ws, lr + 1, 18)
    write_cell(ws, lr + 1, 5, "Must: 필수  /  Should: 권장  /  Nice: 선택",
               font=FONT_SMALL, fill=FILL_GRAY, alignment=ALIGN_LEFT)


def build_bo_req(wb, d):
    ws = wb.create_sheet("BO 요구사항")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [3, 12, 18, 50, 14, 20, 18, 3]):
        set_col_width(ws, col, w)

    section_title(ws, 1, 7, "3. BO (Back Office) 기능 요구사항", height=30)
    header_row(ws, 3,
               ["", "요구사항 ID", "구분", "요구사항 내용", "우선순위", "관련 화면", "비고", ""],
               height=22)

    for i, req in enumerate(d.get("bo_requirements", []), 1):
        row = 3 + i
        pri = req.get("priority", "Must")
        pc, fc = PRIORITY_COLORS.get(pri, (C_GRAY, "000000"))
        bg = fill(pc)
        write_cell(ws, row, 2, req["id"],       font=FONT_SMALL, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 3, req["category"], font=FONT_BODY,  fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 4, req["content"],  font=FONT_BODY,  fill=bg, alignment=ALIGN_LEFT_TOP, border=BORDER_ALL)
        write_cell(ws, row, 5, pri, font=fnt(10, True, fc), fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 6, req.get("screen", ""), font=FONT_SMALL, fill=bg, alignment=ALIGN_LEFT, border=BORDER_ALL)
        write_cell(ws, row, 7, req.get("note", ""),   font=FONT_SMALL, fill=bg, alignment=ALIGN_LEFT, border=BORDER_ALL)
        set_row_height(ws, row, 36)


def build_nonfunc(wb, d):
    ws = wb.create_sheet("비기능 요구사항")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [3, 18, 60, 20, 3]):
        set_col_width(ws, col, w)

    section_title(ws, 1, 4, "4. 비기능 요구사항 / 제약사항", height=30)
    header_row(ws, 3, ["", "항목", "내용", "비고", ""], height=22)
    for i, item in enumerate(d.get("non_func", []), 1):
        data_row(ws, 3 + i, ["", item["category"], item["content"], item.get("note", ""), ""], i % 2 == 0)
        set_row_height(ws, 3 + i, 36)


def generate(data_path: str, output_path: str):
    with open(data_path, encoding="utf-8") as f:
        d = json.load(f)
    wb = Workbook()
    build_cover(wb, d)
    build_analysis(wb, d)
    build_fo_req(wb, d)
    build_bo_req(wb, d)
    build_nonfunc(wb, d)
    wb.save(output_path)
    print(f"[gen_requirements] 저장 완료: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python gen_requirements.py <data.json> <output.xlsx>")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2])
