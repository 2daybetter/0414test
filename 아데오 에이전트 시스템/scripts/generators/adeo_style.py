"""
아데오 공통 스타일 정의 — 모든 Generator가 import해서 사용
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 색상 ────────────────────────────────────────────────────────────────────
C_NAVY       = "1C3557"   # 헤더 배경 (다크 네이비)
C_BLUE       = "2E74B5"   # 서브헤더 배경
C_BLUE_LIGHT = "D6E4F0"   # 2단계 헤더 배경
C_ACCENT     = "E8F3FF"   # 홀수 행 배경
C_WHITE      = "FFFFFF"
C_GRAY       = "F5F7FA"   # 짝수 행 배경
C_BORDER     = "B0C4D8"   # 테두리

# 단계별 색상 (WBS)
STAGE_COLORS = {
    "PM": ("D9E1F2", "1C3557"),  # 배경, 글자
    "AY": ("FCE4D6", "833C00"),
    "DE": ("E2EFDA", "375623"),
    "IM": ("FFF2CC", "7F6000"),
    "TE": ("FCE4D6", "C00000"),
    "OP": ("EAD1DC", "4E1A35"),
}

# ─── 폰트 ────────────────────────────────────────────────────────────────────
def fnt(size=10, bold=False, color="000000", name="맑은 고딕"):
    return Font(name=name, size=size, bold=bold, color=color)

FONT_TITLE   = fnt(16, True, C_WHITE)
FONT_HEADER  = fnt(10, True, C_WHITE)
FONT_SUBHDR  = fnt(10, True, C_NAVY)
FONT_BODY    = fnt(10, False, "1A1A2E")
FONT_SMALL   = fnt(9,  False, "444444")
FONT_BOLD    = fnt(10, True,  "1A1A2E")

# ─── 채우기 ──────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

FILL_NAVY      = fill(C_NAVY)
FILL_BLUE      = fill(C_BLUE)
FILL_BLUE_LT   = fill(C_BLUE_LIGHT)
FILL_ACCENT    = fill(C_ACCENT)
FILL_GRAY      = fill(C_GRAY)
FILL_WHITE     = fill(C_WHITE)

# ─── 정렬 ────────────────────────────────────────────────────────────────────
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left",  vertical="top",    wrap_text=True)

# ─── 테두리 ──────────────────────────────────────────────────────────────────
def _side(style="thin", color=C_BORDER):
    return Side(style=style, color=color)

BORDER_ALL  = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
BORDER_THICK_BOTTOM = Border(
    left=_side(), right=_side(),
    top=_side(), bottom=_side("medium", C_NAVY)
)

# ─── 헬퍼 함수 ───────────────────────────────────────────────────────────────
def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def set_row_height(ws, row_num, height):
    ws.row_dimensions[row_num].height = height

def write_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    return cell

def fill_range(ws, row1, col1, row2, col2, fill_style, border=None):
    """병합 없이 셀 범위에 배경색/테두리만 적용 (대형 헤더 배경 등에 사용)"""
    for r in range(row1, row2+1):
        for c in range(col1, col2+1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_style
            if border:
                cell.border = border

def merge_write(ws, row1, col1, row2, col2, value, font=None, fill=None, alignment=None, border=None):
    # 값과 스타일을 먼저 쓴 후 병합 (openpyxl 3.x 호환)
    cell = ws.cell(row=row1, column=col1)
    cell.value = value
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    ws.merge_cells(start_row=row1, start_column=col1, end_row=row2, end_column=col2)
    return cell

def header_row(ws, row, cols_values: list, bg=FILL_NAVY, fg=FONT_HEADER, height=22):
    """헤더 행 일괄 작성"""
    for col, val in enumerate(cols_values, 1):
        write_cell(ws, row, col, val, font=fg, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
    set_row_height(ws, row, height)

def data_row(ws, row, cols_values: list, alt=False):
    """데이터 행 일괄 작성 (홀짝 교대 배경)"""
    bg = FILL_ACCENT if alt else FILL_GRAY
    for col, val in enumerate(cols_values, 1):
        write_cell(ws, row, col, str(val) if val is not None else "",
                   font=FONT_BODY, fill=bg, alignment=ALIGN_LEFT, border=BORDER_ALL)
    set_row_height(ws, row, 18)

def section_title(ws, row, col_span, title, height=30):
    """섹션 제목 (전체 너비 병합)"""
    merge_write(ws, row, 1, row, col_span, title,
                font=FONT_TITLE, fill=FILL_NAVY, alignment=ALIGN_CENTER, border=BORDER_ALL)
    set_row_height(ws, row, height)

def sub_section(ws, row, col_span, title, height=20):
    """서브섹션 제목"""
    merge_write(ws, row, 1, row, col_span, title,
                font=FONT_SUBHDR, fill=FILL_BLUE_LT, alignment=ALIGN_LEFT, border=BORDER_THICK_BOTTOM)
    set_row_height(ws, row, height)
