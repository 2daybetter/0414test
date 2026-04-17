"""
gen_ia.py — IA 설계서 Excel 생성기 (DE-02)
Usage: python gen_ia.py <data.json> <output.xlsx>
"""
import json, sys, os
from openpyxl import Workbook
sys.path.insert(0, os.path.dirname(__file__))
from adeo_style import *

TYPE_COLORS = {
    "Page":         ("FFFFFF", "000000"),
    "Layer Popup":  ("FFF9C4", "6D4C00"),
    "Popup":        ("FCE4D6", "8B2000"),
    "Link":         ("E8F5E9", "1B5E20"),
}
DEPTH_BG = {
    1: ("1C3557", "FFFFFF"),   # 1depth — 네이비
    2: ("D6E4F0", C_NAVY),     # 2depth — 연파랑
    3: ("F0F7FF", "333333"),   # 3depth — 흰
}


def build_cover(wb, d):
    ws = wb.active
    ws.title = "문서 정보"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [3, 22, 32, 18, 3]):
        set_col_width(ws, col, w)

    fill_range(ws, 1, 1, 3, 5, FILL_NAVY)
    merge_write(ws, 2, 2, 2, 3, "ADEO GROUP", font=fnt(18, True, C_WHITE), fill=FILL_NAVY, alignment=ALIGN_LEFT)
    merge_write(ws, 2, 4, 2, 5, "IA 설계서 DE-02", font=fnt(11, False, "A0C4E8"), fill=FILL_NAVY,
                alignment=Alignment(horizontal="right", vertical="center"))
    for r in range(1, 4):
        set_row_height(ws, r, 22)

    merge_write(ws, 5, 2, 6, 4, d["project_name"],
                font=fnt(18, True, C_NAVY), fill=FILL_WHITE,
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
    set_row_height(ws, 5, 28); set_row_height(ws, 6, 28)
    merge_write(ws, 7, 2, 7, 4, "", fill=fill(C_BLUE)); set_row_height(ws, 7, 4)

    info = [
        ("문서 코드", "DE-02"), ("버전", d.get("version", "v1.0")),
        ("작성일", d.get("doc_date", "")), ("고객사", d.get("client_name", "")),
        ("작성자", d.get("author", "웹기획팀")), ("검토자", d.get("reviewer", "이사님")),
    ]
    for i, (k, v) in enumerate(info, 9):
        write_cell(ws, i, 2, k, font=FONT_BOLD, fill=FILL_BLUE_LT, alignment=ALIGN_CENTER, border=BORDER_ALL)
        merge_write(ws, i, 3, i, 4, v, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_LEFT, border=BORDER_ALL)
        set_row_height(ws, i, 20)


def build_ia_sheet(wb, d, side: str):
    """FO 또는 BO IA 시트 생성"""
    ws = wb.create_sheet(f"{side} IA")
    ws.sheet_view.showGridLines = False

    # 열 너비
    col_cfg = [
        ("A", 3), ("B", 14), ("C", 6), ("D", 18), ("E", 16), ("F", 16),
        ("G", 14), ("H", 10), ("I", 8), ("J", 30), ("K", 28),
        ("L", 28), ("M", 22), ("N", 3),
    ]
    for col, w in col_cfg:
        set_col_width(ws, col, w)

    title_label = "FO (Front Office)" if side == "FO" else "BO (Back Office)"
    doc_code = "DE-02"
    section_title(ws, 1, 13, f"{title_label} IA 설계서  {doc_code}", height=32)

    # 헤더
    headers = ["", "IA_ID", "Depth", "Depth1", "Depth2", "Depth3",
               "Type", "화면ID", "DB연동", "기능 정의", "URL",
               "SEO Title", "SEO Description", ""]
    if side == "BO":
        headers[-2] = "비고"   # BO는 SEO 대신 비고
    header_row(ws, 3, headers, height=22)

    items = d.get(f"{side.lower()}_ia", [])
    prev_d1 = None
    row = 4
    for item in items:
        d1 = item.get("depth1", "")
        d2 = item.get("depth2", "")
        d3 = item.get("depth3", "")
        typ = item.get("type", "Page")
        depth = 3 if d3 else (2 if d2 else 1)

        # 1depth 구분선
        if d1 != prev_d1:
            sc, fc = DEPTH_BG[1]
            merge_write(ws, row, 2, row, 13, f"  {d1}",
                        font=fnt(10, True, fc), fill=fill(sc),
                        alignment=ALIGN_LEFT, border=BORDER_ALL)
            set_row_height(ws, row, 20)
            row += 1
            prev_d1 = d1

        # 타입 색상
        tc, tfc = TYPE_COLORS.get(typ, (C_WHITE, "000000"))
        bg = fill(tc)

        # 깊이별 배경
        dc, dfc = DEPTH_BG[depth]
        depth_bg = fill(dc)

        write_cell(ws, row, 2, item.get("ia_id", ""),  font=FONT_SMALL, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 3, f"{depth}depth",        font=FONT_SMALL, fill=depth_bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 4, d1,  font=fnt(10, depth == 1, dfc if depth == 1 else "000000"),
                   fill=depth_bg if depth == 1 else bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 5, d2,  font=FONT_BODY, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 6, d3,  font=FONT_BODY, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 7, typ, font=fnt(9, True, tfc), fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 8, item.get("screen_id", ""), font=FONT_SMALL, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 9, item.get("db", "N"),       font=FONT_BODY,  fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 10, item.get("function", ""), font=FONT_BODY,  fill=bg, alignment=ALIGN_LEFT,   border=BORDER_ALL)
        write_cell(ws, row, 11, item.get("url", ""),      font=fnt(9), fill=bg, alignment=ALIGN_LEFT, border=BORDER_ALL)
        write_cell(ws, row, 12, item.get("seo_title", item.get("note", "")), font=FONT_SMALL, fill=bg, alignment=ALIGN_LEFT, border=BORDER_ALL)
        write_cell(ws, row, 13, item.get("seo_desc", ""),  font=FONT_SMALL, fill=bg, alignment=ALIGN_LEFT_TOP, border=BORDER_ALL)
        set_row_height(ws, row, 24)
        row += 1

    # 범례
    row += 1
    sub_section(ws, row, 13, "  Type 범례", height=18)
    row += 1
    for j, (typ, (tc, tfc)) in enumerate(TYPE_COLORS.items(), 2):
        write_cell(ws, row, j, typ, font=fnt(9, True, tfc), fill=fill(tc), alignment=ALIGN_CENTER, border=BORDER_ALL)
    set_row_height(ws, row, 18)


def build_summary_sheet(wb, d):
    ws = wb.create_sheet("화면 수 집계")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [3, 12, 12, 14, 14, 14, 3]):
        set_col_width(ws, col, w)

    section_title(ws, 1, 6, "전체 화면 수 집계표", height=30)
    header_row(ws, 3, ["", "구분", "Page", "Layer Popup", "Popup", "합계", ""], height=22)

    summary = d.get("summary", {})
    rows = [("FO", "fo"), ("BO", "bo"), ("합계", "total")]
    for i, (label, key) in enumerate(rows, 1):
        s = summary.get(key, {})
        is_total = (key == "total")
        bg = FILL_NAVY if is_total else (FILL_ACCENT if i % 2 == 0 else FILL_GRAY)
        ft = fnt(10, is_total, C_WHITE if is_total else "000000")
        row = 3 + i
        write_cell(ws, row, 2, label,               font=ft, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 3, s.get("page", 0),    font=ft, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 4, s.get("layer", 0),   font=ft, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 5, s.get("popup", 0),   font=ft, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 6, s.get("total", 0),   font=ft, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        set_row_height(ws, row, 22)


def generate(data_path: str, output_path: str):
    with open(data_path, encoding="utf-8") as f:
        d = json.load(f)
    wb = Workbook()
    build_cover(wb, d)
    build_ia_sheet(wb, d, "FO")
    build_ia_sheet(wb, d, "BO")
    build_summary_sheet(wb, d)
    wb.save(output_path)
    print(f"[gen_ia] 저장 완료: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python gen_ia.py <data.json> <output.xlsx>")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2])
