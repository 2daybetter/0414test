"""
gen_wbs.py — WBS / 킥오프 보고서 Excel 생성기 (PM-01, PM-02, PM-03)
Usage: python gen_wbs.py <data.json> <output.xlsx>
"""
import json, sys, os
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
sys.path.insert(0, os.path.dirname(__file__))
from adeo_style import *


def _weeks(start, pct, total_days):
    days = int(total_days * pct / 100)
    return start, start + timedelta(days=days - 1)


def build_cover_sheet(wb, d):
    ws = wb.active
    ws.title = "표지"
    ws.sheet_view.showGridLines = False

    # 열 너비
    for col, w in zip("ABCDEFG", [3, 18, 28, 18, 14, 14, 3]):
        set_col_width(ws, col, w)

    # 상단 로고 영역 — 먼저 배경색 채우고(병합 없이), 텍스트 셀만 병합
    fill_range(ws, 1, 1, 3, 7, FILL_NAVY)
    merge_write(ws, 2, 2, 2, 3, "ADEO GROUP", font=fnt(22, True, C_WHITE), fill=FILL_NAVY, alignment=ALIGN_LEFT)
    merge_write(ws, 2, 4, 2, 7, "프로젝트 사업수행계획서", font=fnt(14, False, "A0C4E8"), fill=FILL_NAVY, alignment=Alignment(horizontal="right", vertical="center"))
    for r in range(1, 4):
        set_row_height(ws, r, 22)

    # 프로젝트 제목
    merge_write(ws, 5, 2, 6, 6, d["project_name"],
                font=fnt(20, True, C_NAVY), fill=FILL_WHITE,
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
    set_row_height(ws, 5, 30); set_row_height(ws, 6, 30)

    # 구분선
    merge_write(ws, 7, 2, 7, 6, "", fill=fill(C_BLUE))
    set_row_height(ws, 7, 4)

    # 기본 정보 테이블
    info_rows = [
        ("문서 코드", "PM-01"),
        ("버전", d.get("version", "v1.0")),
        ("작성일", d.get("doc_date", datetime.today().strftime("%Y-%m-%d"))),
        ("고객사", d["client_name"]),
        ("계약 기간", f"{d['start_date']} ~ {d['end_date']}"),
        ("작성자", d.get("author", "PM")),
        ("검토자", d.get("reviewer", "이사님")),
        ("승인자", d.get("approver", "이사님")),
    ]
    for i, (k, v) in enumerate(info_rows, 9):
        write_cell(ws, i, 2, k, font=FONT_BOLD, fill=FILL_BLUE_LT, alignment=ALIGN_CENTER, border=BORDER_ALL)
        merge_write(ws, i, 3, i, 6, v, font=FONT_BODY, fill=FILL_WHITE, alignment=ALIGN_LEFT, border=BORDER_ALL)
        set_row_height(ws, i, 20)

    # 하단 여백
    merge_write(ws, 20, 2, 21, 6, "본 문서는 아데오 그룹 내부 산출물로, 무단 복제 및 배포를 금합니다.",
                font=FONT_SMALL, fill=FILL_GRAY, alignment=ALIGN_CENTER)
    set_row_height(ws, 20, 16); set_row_height(ws, 21, 16)


def build_wbs_sheet(wb, d):
    ws = wb.create_sheet("WBS")
    ws.sheet_view.showGridLines = False

    COLS = ["코드", "산출물 코드", "산출물명", "담당팀", "시작일", "종료일", "진행률", "비고"]
    widths = [3, 10, 12, 22, 16, 12, 12, 10, 12, 3]
    for col, w in zip("ABCDEFGHIJ", widths):
        set_col_width(ws, col, w)

    # 제목
    section_title(ws, 1, 9, "WBS (Work Breakdown Structure)   PM-03", height=32)

    # 마일스톤 요약
    sub_section(ws, 3, 9, "  마일스톤 일정", height=22)
    header_row(ws, 4, ["단계", "단계명", "가중치", "시작일", "종료일", "담당팀", "비고", "", ""],
               height=20)

    stages = d.get("stages", [])
    for i, st in enumerate(stages):
        alt = (i % 2 == 0)
        bg = FILL_ACCENT if alt else FILL_GRAY
        sc, fc = STAGE_COLORS.get(st["code"], (C_GRAY, "000000"))
        row = 5 + i
        write_cell(ws, row, 2, st["code"],    font=fnt(10, True, fc), fill=fill(sc), alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 3, st["name"],    font=fnt(10, True, fc), fill=fill(sc), alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 4, st["weight"],  font=FONT_BODY, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 5, st["start"],   font=FONT_BODY, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 6, st["end"],     font=FONT_BODY, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        write_cell(ws, row, 7, st.get("team", ""), font=FONT_BODY, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
        merge_write(ws, row, 8, row, 9, st.get("note", ""), font=FONT_BODY, fill=bg, alignment=ALIGN_LEFT, border=BORDER_ALL)
        set_row_height(ws, row, 20)

    # WBS 상세
    detail_start = 6 + len(stages)
    sub_section(ws, detail_start, 9, "  WBS 상세", height=22)
    header_row(ws, detail_start + 1,
               ["", "코드", "산출물명", "담당팀", "시작일", "종료일", "진행률", "비고", ""],
               height=20)

    row = detail_start + 2
    for st in d.get("wbs_detail", []):
        # 단계 구분 행
        sc, fc = STAGE_COLORS.get(st["stage_code"], (C_GRAY, "000000"))
        label = f"[{st['stage_code']}] {st['stage_name']} — {st['weight']}"
        merge_write(ws, row, 2, row, 9, label,
                    font=fnt(10, True, fc), fill=fill(sc),
                    alignment=ALIGN_LEFT, border=BORDER_ALL)
        set_row_height(ws, row, 20)
        row += 1
        for j, item in enumerate(st["items"]):
            alt = (j % 2 == 0)
            bg = FILL_ACCENT if alt else FILL_GRAY
            write_cell(ws, row, 2, item["code"],     font=FONT_SMALL, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
            write_cell(ws, row, 3, item["doc_name"], font=FONT_BODY,  fill=bg, alignment=ALIGN_LEFT,   border=BORDER_ALL)
            write_cell(ws, row, 4, item["team"],     font=FONT_BODY,  fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
            write_cell(ws, row, 5, item["start"],    font=FONT_BODY,  fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
            write_cell(ws, row, 6, item["end"],      font=FONT_BODY,  fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
            write_cell(ws, row, 7, item.get("progress", "0%"), font=FONT_BODY, fill=bg, alignment=ALIGN_CENTER, border=BORDER_ALL)
            merge_write(ws, row, 8, row, 9, item.get("note", ""), font=FONT_SMALL, fill=bg, alignment=ALIGN_LEFT, border=BORDER_ALL)
            set_row_height(ws, row, 18)
            row += 1


def build_risk_sheet(wb, d):
    ws = wb.create_sheet("리스크 관리")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [3, 8, 30, 16, 12, 18, 30, 3]):
        set_col_width(ws, col, w)

    section_title(ws, 1, 7, "리스크 관리 계획", height=30)
    sub_section(ws, 3, 7, "  리스크 목록", height=20)
    header_row(ws, 4, ["", "번호", "리스크 항목", "발생 가능성", "영향도", "담당자", "대응 방안", ""], height=22)

    for i, r in enumerate(d.get("risks", []), 1):
        alt = (i % 2 == 0)
        data_row(ws, 4 + i,
                 ["", f"R-{i:02d}", r["item"], r["probability"], r["impact"], r.get("owner", "PM"), r["action"], ""], alt)


def build_org_sheet(wb, d):
    ws = wb.create_sheet("조직도")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [3, 16, 18, 14, 14, 3]):
        set_col_width(ws, col, w)

    section_title(ws, 1, 5, "프로젝트 조직 및 인력 계획 (PM-02)", height=30)

    sub_section(ws, 3, 5, "  프로젝트 조직", height=20)
    header_row(ws, 4, ["", "역할", "담당자", "소속", "담당 업무", ""], height=22)
    for i, m in enumerate(d.get("members", []), 1):
        data_row(ws, 4 + i, ["", m["role"], m["name"], m.get("dept", "아데오"), m.get("task", ""), ""], i % 2 == 0)

    # 단계별 투입 인력
    hr = 6 + len(d.get("members", []))
    sub_section(ws, hr, 5, "  단계별 투입 인력 계획", height=20)
    stage_codes = [s["code"] for s in d.get("stages", [])]
    header_row(ws, hr + 1, ["", "팀"] + stage_codes + [""], height=22)
    for j, t in enumerate(d.get("resource_plan", []), 1):
        row = hr + 1 + j
        vals = ["", t["team"]] + [t.get(s, "-") for s in stage_codes] + [""]
        data_row(ws, row, vals, j % 2 == 0)


def generate(data_path: str, output_path: str):
    with open(data_path, encoding="utf-8") as f:
        d = json.load(f)

    wb = Workbook()
    build_cover_sheet(wb, d)
    build_wbs_sheet(wb, d)
    build_risk_sheet(wb, d)
    build_org_sheet(wb, d)

    wb.save(output_path)
    print(f"[gen_wbs] 저장 완료: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python gen_wbs.py <data.json> <output.xlsx>")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2])
